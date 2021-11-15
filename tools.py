import os
import re
import shutil
import openpyxl

def __find_all_file(base):
    file_list = []
    for root, ds, fs in os.walk(base):
        for f in fs:
            full_name = os.path.join(root, f)
            file_list.append(full_name)
    return file_list    
            

def split_log(regex,path):
    string_list = []
    temp = []
    file_list = __find_all_file(path)
    pattern = re.compile(r"(?<=_)\d{4}")
    i = 1
    for index, file_name in enumerate(file_list):
        index = round(index/3)
        if(file_name[-3:]=="log"):
            number_list = pattern.findall(file_name)
            with open(file_name,"r",encoding='utf-8', errors='ignore') as file:
                lines = file.readlines()
                temp = lines[(i-1)*1007:1007*i] 
                i = i+1
            with open(file_name,"w",encoding='utf-8', errors='ignore') as file:
                file.writelines(temp)



def auto_create_folder(before_name,after_name):
    name_list = []
    with open("./config","r",encoding='utf-8') as config_file:
        name_list = config_file.readlines()
    for name in name_list:
        now_file_name = before_name+name.strip("\n")+after_name
        shutil.copytree("C:/Users/001496-liteng/Desktop/Bridgestone/STEF/1115/No.203_0920",f"C:/Users/001496-liteng/Desktop/Bridgestone/STEF/1115/{now_file_name}")
        
        

def auto_rename_file(path,before_name,tsv_before_name):
    string_list = []
    temp = []
    file_list = __find_all_file(path)
    number_pattern = re.compile(r"(?<=_)\d{4}")
    excel_name_pattern = re.compile(r"(?<=_)\d{4}\(.+\)")
    for root, ds, fs in os.walk(path):
        for f in fs:
            number=number_pattern.findall(root)
            excel_name = excel_name_pattern.findall(root)
            if(f[-4:] == "xlsx"):
               os.rename(os.path.join(root,f),os.path.join(root,before_name+excel_name[0]+".xlsx"))
            elif(f[-3:] == "tsv"):
                all_line = []
                with open(os.path.join(root,f),"r") as tsv:
                    all_line = tsv.readlines()
                with open(os.path.join(root,f),"w") as tsv:
                    for line in all_line:
                        a = re.sub("0920",number[0],line)
                        tsv.writelines(a)
                os.rename(os.path.join(root,f),os.path.join(root,tsv_before_name+number[0]+".tsv"))
                #shutil.copyfile(os.path.join(root,tsv_before_name+number[0]+".tsv"),os.path.join("N:/hulft/recv",tsv_before_name+number[0]+".tsv"))
      
      
def repalce_all_excel(path):
    number_pattern = re.compile(r"(?<=_)\d{4}")
    file_list = __find_all_file(path)
    for file_name in file_list:
        number = number_pattern.findall(file_name)
        if(file_name[-4:]=="xlsx"):    
            wb = openpyxl.load_workbook(file_name)
            ws = wb["異常系"]
            i = 0
            for r in range(1,ws.max_row+1):
                for c in range(1,ws.max_column+1):
                    s = ws.cell(r,c).value
                    if not isinstance(s,str):
                        s = str(s)
                    if s != None and "1017" in s: 
                        ws.cell(r,c).value = s.replace("1017",number[0]) 
                        print("row {} col {} : {}".format(r,c,s))
                        i += 1
                wb.save(file_name)
                print("{} cells updated".format(i))


def paste_log_to_excel(path):
    number_pattern = re.compile(r"(?<=_)\d{4}")
    file_list = __find_all_file(path)
        
    for file_name in file_list:
        if file_name[-4:]=="xlsx":    
            number = number_pattern.findall(file_name)
            log_file_path = "N:/appl/logs/DEB.log"
            line_number = 0
            data = []
            with open("N:/appl/logs/DEB.log","r",encoding='utf-8') as log_file:
                lines = log_file.readlines()
                for index, line in enumerate(lines):
                    key_word = f"項目NO[](komokuNo):[{number[0]}]"
                    if(key_word in line ):
                        line_number = index
                        data = lines[line_number-11:line_number+1]
            wb = openpyxl.load_workbook(file_name)
            ws = wb["異常系"]
            for index,line in enumerate(data):
                ws[f"B{index+26}"] = line
                ws["B25"] = ""
            wb.save(file_name)     
            
#paste_log_to_excel("C:/Users/001496-liteng/Desktop/Bridgestone/STEF/1115")
#auto_create_folder("No.205_","")
#repalce_all_excel("C:/Users/001496-liteng/Desktop/Bridgestone/STEF/1115")
#auto_rename_file("C:/Users/001496-liteng/Desktop/Bridgestone/STEF/1115","エビデンス_CMCA740B01_得意先マスタ申請WF結果更新(一括)（更新）チェックNo.205_","cmca740btest_no205_")
split_log("",path="C:/Users/001496-liteng/Desktop/Bridgestone/STEF/1115")