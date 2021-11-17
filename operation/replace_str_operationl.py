from os import path
import openpyxl,re
from builder.file_looper import FileLooper
from operation.operation import Operation



class ReplaceStrOperation(Operation):
    def __init__(self,regex,text_after_replacement,file_path = "",priority = 0):
        self.file_path = file_path
        self.priority = priority
        self.regex = regex 
        self.text_after_replacement = text_after_replacement
    
    def execute(self):
        pattern = re.compile(self.regex)       
        if self.file_path[-4:]=="xlsx":    
            wb = openpyxl.load_workbook(self.file_path)
            for ws in wb:
                i = 0
                for r in range(1,ws.max_row+1):
                    for c in range(1,ws.max_column+1):
                        s = str(ws.cell(r,c).value)
                        if s != None and pattern.match(s): 
                            ws.cell(r,c).value = pattern.sub(s,self.text_after_replacement)
                            print("row {} col {} : {}".format(r,c,s))
                            i += 1
                wb.save(self.file_path)
                print("{} cells updated".format(i))